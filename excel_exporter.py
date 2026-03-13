"""
Excel 日报导出
==============
将 ReportBuilder 实例的数据导出为格式化 Excel 文件。
包含：概览卡片、统计表、饼图、各板块工单表格、超R柱状图。
"""
import os
import re
import logging
import tempfile
from pathlib import Path
from datetime import datetime

from config import (
    SUPER_R_THRESHOLD, STATUS_SOLVED, STATUS_CLOSED, ts_to_str,
)

logger = logging.getLogger(__name__)


def build_excel(builder, filepath=None):
    """
    生成增强版 Excel 日报（概览卡片、统计图、各板块工单表格）。
    :param builder: ReportBuilder 实例
    :param filepath: 输出路径，None 则使用默认路径
    :return: 保存路径字符串
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils import get_column_letter

    filepath = filepath or builder._output_path("xlsx")
    wb = Workbook()
    ws = wb.active

    _illegal_chars_re = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

    def safe(val):
        return _illegal_chars_re.sub('', val) if isinstance(val, str) else val

    ws.title = "VIP客服日报"

    # ---- 样式 ----
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    section_font = Font(bold=True, size=11, color="2F5496")
    normal_font = Font(size=10)
    bold_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(bold=True, size=10, color="9C0006")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_font = Font(size=10, color="9C6500")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(size=10, color="006100")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row = 1
    date_str_cn = builder.report_date.strftime("%Y年%m月%d日")

    # ---- 标题 ----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=safe(f"VIP客服日报 - {date_str_cn}")).font = title_font
    row += 1

    # 警告信息
    if builder.errors:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=safe(f"[!] 数据获取不完整：{'、'.join(builder.errors)}"))
        cell.font = red_font
        cell.fill = red_fill
        row += 1
    if builder.dedup_removed > 0:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=safe(f"[i] 已自动去除 {builder.dedup_removed} 条重复工单")).font = normal_font
        row += 1

    row += 1

    # ---- 概览统计卡片 ----
    overview_headers = ["当日工单数", "待跟进工单", "总会话量", "超R工单", "预流失", "预投诉"]
    overview_values = [
        len(builder.daily_tickets), len(builder.pending_tickets), builder.total_sessions,
        len(builder._super_r), len(builder._pre_churn), len(builder._pre_complaint),
    ]
    for col, h in enumerate(overview_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center
    row += 1
    for col, v in enumerate(overview_values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = Font(bold=True, size=14, color="2F5496")
        cell.border = thin_border
        cell.alignment = center
    # 趋势行
    if builder.trend_data:
        row += 1
        trend_keys = ["prev_daily_count", "prev_pending_count", "prev_total_sessions",
                      "prev_super_r_count", "prev_pre_churn_count", "prev_pre_complaint_count"]
        for col, (val, key) in enumerate(zip(overview_values, trend_keys), 1):
            prev = builder.trend_data.get(key)
            if prev is not None:
                diff = val - prev
                mark = f"+{diff}" if diff > 0 else str(diff) if diff < 0 else "持平"
                cell = ws.cell(row=row, column=col, value=safe(f"较昨日 {mark}"))
                cell.alignment = center
                if diff > 0:
                    cell.font = red_font
                elif diff < 0:
                    cell.font = green_font
                else:
                    cell.font = normal_font
    row += 2

    # ---- 统计表 ----
    categories, cat_alarm, cat_dev = builder._cat_stats

    stat_headers = ["问题类型", "工单/会话数", "客服处理", "运营/研发介入"]
    for col, h in enumerate(stat_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center
    row += 1

    chart_start_row = row
    for cat in categories:
        a = cat_alarm.get(cat, 0)
        d = cat_dev.get(cat, 0)
        if a == 0:
            continue
        vals = [cat, a, a - d, d]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=safe(v))
            cell.font = normal_font
            cell.border = thin_border
            if col > 1:
                cell.alignment = center
        row += 1
    chart_end_row = row - 1

    # 合计行
    total_a = sum(cat_alarm.values())
    total_d = sum(cat_dev.values())
    for col, v in enumerate(["合计", total_a, total_a - total_d, total_d], 1):
        cell = ws.cell(row=row, column=col, value=safe(v))
        cell.font = bold_font
        cell.border = thin_border
        cell.fill = light_fill
        if col > 1:
            cell.alignment = center
    row += 1

    ws.cell(row=row, column=1, value=safe(f"总会话量：{builder.total_sessions}")).font = bold_font
    row += 1

    # ---- 饼图：问题类型分布 ----
    if chart_end_row >= chart_start_row:
        pie = PieChart()
        pie.title = "问题类型分布"
        pie.width = 18
        pie.height = 12
        labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
        data = Reference(ws, min_col=2, min_row=chart_start_row - 1, max_row=chart_end_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.style = 10
        ws.add_chart(pie, f"F{chart_start_row}")

    row += 16  # 为图表留空间

    # ---- 七个板块（详细工单表格） ----

    def _write_section_table(ws, row, section_title, orders,
                             show_amount=False, show_resolved=False):
        """写一个板块的工单表格"""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=safe(section_title))
        cell.font = section_font
        row += 1

        if not orders:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=1, value="无").font = normal_font
            return row + 1

        # 表头
        cols = ["序号", "发起人", "工单内容", "创建时间", "更新时间", "受理人"]
        if show_amount:
            cols.append("累充金额")
        if show_resolved:
            cols.append("状态")
        for col, h in enumerate(cols, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center
        row += 1

        for i, t in enumerate(orders, 1):
            values = [
                i,
                t.get("_creator", ""),
                builder._summarize(t),
                ts_to_str(t.get("_create_time", 0)),
                ts_to_str(t.get("_update_time", 0)),
                t.get("_handler", ""),
            ]
            if show_amount:
                amt = t["_recharge"]
                values.append(f"{amt/10000:.1f}W" if amt >= 10000 else f"{amt:.0f}")
            if show_resolved:
                resolved = t["_status"] in (STATUS_SOLVED, STATUS_CLOSED)
                values.append("已解决" if resolved else "未解决")

            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=safe(v))
                cell.font = normal_font
                cell.border = thin_border
                if col == 3:  # 工单内容列自动换行
                    cell.alignment = wrap
                else:
                    cell.alignment = Alignment(vertical="center")

            # 条件格式：超R行红色高亮
            if show_amount and t["_recharge"] >= SUPER_R_THRESHOLD:
                for col in range(1, len(values) + 1):
                    ws.cell(row=row, column=col).fill = red_fill
                    ws.cell(row=row, column=col).font = red_font

            # 条件格式：未解决标黄
            if show_resolved and t["_status"] not in (STATUS_SOLVED, STATUS_CLOSED):
                for col in range(1, len(values) + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill
                    if col == len(values):
                        ws.cell(row=row, column=col).font = yellow_font

            # 条件格式：已解决标绿
            if show_resolved and t["_status"] in (STATUS_SOLVED, STATUS_CLOSED):
                status_col = len(values)
                ws.cell(row=row, column=status_col).fill = green_fill
                ws.cell(row=row, column=status_col).font = green_font

            row += 1

        return row + 1

    # 二、待跟进
    row = _write_section_table(ws, row, f"二、待跟进问题总计 (共{len(builder._pending_dev)}条)", builder._pending_dev)

    # 三、未回访
    row = _write_section_table(ws, row, f"三、客服未回访/未跟进 (共{len(builder._unvisited)}条)", builder._unvisited)

    # 四、超R
    row = _write_section_table(ws, row, f"四、超R反馈问题 (共{len(builder._super_r)}条)", builder._super_r, show_amount=True)

    # 五、预流失
    row = _write_section_table(ws, row, f"五、预流失报备 (共{len(builder._pre_churn)}条)", builder._pre_churn)

    # 六、预投诉
    row = _write_section_table(ws, row, f"六、预投诉报备 (共{len(builder._pre_complaint)}条)", builder._pre_complaint, show_resolved=True)

    # 七、其他
    row = _write_section_table(ws, row, f"七、其他VIP用户反馈 (共{len(builder._other)}条)", builder._other, show_resolved=True)

    # ---- 列宽 ----
    ws.column_dimensions["A"].width = 6   # 序号
    ws.column_dimensions["B"].width = 14  # 发起人
    ws.column_dimensions["C"].width = 55  # 工单内容
    ws.column_dimensions["D"].width = 18  # 创建时间
    ws.column_dimensions["E"].width = 18  # 更新时间
    ws.column_dimensions["F"].width = 16  # 受理人
    ws.column_dimensions["G"].width = 12  # 累充金额/状态
    ws.column_dimensions["H"].width = 10

    # ---- 冻结表头 ----
    ws.freeze_panes = "A4"

    # ---- 超R金额柱状图（独立Sheet）----
    super_r = builder._super_r
    if super_r:
        ws2 = wb.create_sheet("超R金额分布")
        ws2.cell(row=1, column=1, value="玩家").font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2, value="累充金额(万)").font = header_font
        ws2.cell(row=1, column=2).fill = header_fill
        for i, t in enumerate(super_r, 2):
            ws2.cell(row=i, column=1, value=safe(t.get("_creator", ""))).font = normal_font
            ws2.cell(row=i, column=2, value=round(t["_recharge"] / 10000, 1)).font = normal_font
        bar = BarChart()
        bar.title = "超R玩家累充金额分布"
        bar.y_axis.title = "金额（万元）"
        bar.x_axis.title = "玩家"
        bar.width = 24
        bar.height = 14
        cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(super_r))
        vals = Reference(ws2, min_col=2, min_row=1, max_row=1 + len(super_r))
        bar.add_data(vals, titles_from_data=True)
        bar.set_categories(cats)
        bar.style = 10
        ws2.add_chart(bar, "D1")
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 15

    # 写入文件（先写临时文件再替换，避免文件锁问题）
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=Path(filepath).parent)
    try:
        os.close(tmp_fd)
        wb.save(tmp_path)
        target = Path(filepath)
        if target.exists():
            try:
                target.unlink()
            except PermissionError:
                ts = datetime.now().strftime("%H%M%S")
                filepath = str(target.with_stem(target.stem + f"_{ts}"))
        Path(tmp_path).replace(filepath)
    except Exception:
        Path(tmp_path).unlink(missing_ok=True)
        raise

    logger.info(f"Excel日报已保存: {filepath}")
    return str(filepath)
