"""
客服日报生成器
==============
将七鱼 OpenAPI 返回的工单数据处理为标准化日报。

日报结构：
  顶部：问题类型统计表
  1. 重大事件报备（批量问题）
  2. 待跟进问题总计
  3. 客服未回访/未跟进问题总计
  4. 超R反馈问题（10w+）
  5. 预流失报备
  6. 预投诉报备
  7. 其他VIP用户反馈

增强功能：
  - 工单去重（按用户+时间窗口）
  - 增强 Excel（条件格式、图表、冻结表头）
  - PDF 导出
  - 趋势对比
  - LLM 智能分类/摘要
"""
import os
import re
import logging
import tempfile
from datetime import datetime
from collections import Counter, defaultdict
from pathlib import Path

from config import (
    SUPER_R_THRESHOLD, ISSUE_KEYWORDS,
    OUTPUT_DIR, HANDLER_DEV_KEYWORD, HANDLER_SYSTEM_KEYWORD,
    DEV_TRANSFER_KEYWORD,
    CF_ISSUE_TYPE, CF_RECHARGE, CF_ISSUE_SELECT,
    STATUS_SOLVED, STATUS_CLOSED,
    ts_to_str, LLM_ENABLED,
)

logger = logging.getLogger(__name__)


# ==================== 工单字段提取工具 ====================

def get_custom_field(custom_list, field_name, default=""):
    """从自定义字段列表中提取值"""
    if isinstance(custom_list, list):
        for item in custom_list:
            name = item.get("name", "") or item.get("key", "")
            if name == field_name:
                return item.get("value", default)
    elif isinstance(custom_list, dict):
        return custom_list.get(field_name, default)
    return default


def parse_amount(raw):
    """解析金额字符串为数值（元）"""
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        raw = raw.replace(",", "").replace("，", "").strip()
        m = re.match(r"([\d.]+)\s*[万wW]", raw)
        if m:
            return float(m.group(1)) * 10000
        try:
            return float(raw)
        except ValueError:
            return 0
    return 0


def enrich_ticket_fields(t):
    """为工单提取常用便捷属性，以 _ 前缀存储"""
    t.setdefault("_handler", "")
    t.setdefault("_has_dev_transfer", False)
    t["_id"] = t.get("id", "")
    t["_title"] = t.get("title", "")
    t["_content"] = t.get("content", "") or t.get("title", "")

    custom = t.get("custom", [])

    # 发起人优先级：角色名(自定义字段) > 标题中提取 > crmUserName > userName > creator
    role_name = get_custom_field(custom, "角色名")
    title_name = ""
    title = t.get("title", "")
    m = re.search(r"】(.+?)[:：]", title)
    if m:
        title_name = m.group(1).strip()
    t["_creator"] = (
        role_name or title_name
        or t.get("crmUserName", "") or t.get("userName", "") or t.get("creator", "")
    )

    t["_create_time"] = t.get("createTime", 0)
    t["_update_time"] = t.get("updateTime", 0) or t.get("createTime", 0)
    t["_status"] = t.get("status", -1)

    t["_issue_select"] = get_custom_field(custom, CF_ISSUE_SELECT)
    t["_issue_type"] = get_custom_field(custom, CF_ISSUE_TYPE)
    t["_recharge"] = parse_amount(get_custom_field(custom, CF_RECHARGE, "0"))


def classify_ticket(t):
    """根据关键词（或AI分类结果）对工单进行问题分类"""
    # 优先使用AI分类结果
    ai_cat = t.get("_ai_category")
    if ai_cat:
        return ai_cat

    text = (t.get("_content") or "") + (t.get("_title") or "")
    for category, keywords in ISSUE_KEYWORDS.items():
        if any(kw in text for kw in keywords):
            return category
    return "其他问题"


# ==================== 工单去重 ====================

def dedup_tickets(tickets, time_window_ms=3600000):
    """
    工单去重：同一用户在时间窗口内的重复工单只保留最新一条。
    :param tickets: 工单列表
    :param time_window_ms: 时间窗口（毫秒），默认1小时
    :return: 去重后的工单列表
    """
    if not tickets:
        return tickets

    # 按创建时间倒序
    sorted_tickets = sorted(tickets, key=lambda t: t.get("_create_time", 0), reverse=True)

    seen = {}  # {creator: [(ticket, create_time), ...]}
    result = []
    dup_count = 0

    for t in sorted_tickets:
        creator = t.get("_creator", "")
        ct = t.get("_create_time", 0)
        content_key = re.sub(r'\s+', '', t.get("_title", ""))[:30]
        dedup_key = f"{creator}|{content_key}"

        if dedup_key in seen:
            prev_time = seen[dedup_key]
            if abs(ct - prev_time) < time_window_ms:
                dup_count += 1
                logger.debug(f"去重: #{t.get('_id')} 与已有工单重复 (creator={creator})")
                continue

        seen[dedup_key] = ct
        result.append(t)

    if dup_count > 0:
        logger.info(f"工单去重: {len(tickets)} → {len(result)} 条 (去除 {dup_count} 条重复)")

    return result


# ==================== 统计计算 ====================

def compute_category_stats(session_data, daily_tickets):
    """
    计算问题类型统计数据（从会话监控数据中提取）。
    运营/研发介入定义：工单日志中存在转交企业微信-飞鱼科技的记录。
    返回 (categories, cat_alarm, cat_dev) 三元组。
    """
    categories = list(ISSUE_KEYWORDS.keys()) + ["其他问题"]

    # 运营介入判定：工单日志中有转交飞鱼科技的记录
    dev_creators = {
        t["_creator"] for t in daily_tickets
        if t.get("_has_dev_transfer", False)
    }

    cat_alarm = Counter()
    cat_dev = Counter()
    for sess in session_data:
        content = sess.get("content", "") or sess.get("message", "") or ""
        visitor = sess.get("visitorName", "") or sess.get("userId", "")
        matched = False
        for category, keywords in ISSUE_KEYWORDS.items():
            if any(kw in content for kw in keywords):
                cat_alarm[category] += 1
                if visitor in dev_creators:
                    cat_dev[category] += 1
                matched = True
                break
        if not matched:
            cat_alarm["其他问题"] += 1
            if visitor in dev_creators:
                cat_dev["其他问题"] += 1

    return categories, cat_alarm, cat_dev


def compute_ticket_category_stats(daily_tickets):
    """
    基于工单数据计算问题类型统计（当会话数据不可用时的备用方案）。
    运营/研发介入定义：工单日志中存在转交企业微信-飞鱼科技的记录。
    """
    categories = list(ISSUE_KEYWORDS.keys()) + ["其他问题"]
    cat_count = Counter()
    cat_dev = Counter()

    for t in daily_tickets:
        cat = classify_ticket(t)
        cat_count[cat] += 1
        if t.get("_has_dev_transfer", False):
            cat_dev[cat] += 1

    return categories, cat_count, cat_dev


# ==================== 日报构建器 ====================

class ReportBuilder:
    """客服日报构建器"""

    def __init__(self, daily_tickets=None, pending_tickets=None,
                 total_sessions=0, report_date=None, session_data=None,
                 trend_data=None, errors=None):
        """
        :param daily_tickets:   当日工单列表
        :param pending_tickets: 近30天待跟进工单列表
        :param total_sessions:  总会话量
        :param report_date:     日报日期
        :param session_data:    会话监控数据
        :param trend_data:      趋势对比数据 {"prev_daily_count": int, "prev_pending_count": int, ...}
        :param errors:          数据获取错误列表
        """
        self.daily_tickets = daily_tickets or []
        self.pending_tickets = pending_tickets or []
        self.total_sessions = total_sessions
        self.report_date = report_date or datetime.now()
        self.session_data = session_data or []
        self.trend_data = trend_data or {}
        self.errors = errors or []

        # 预处理：提取每条工单的便捷属性
        for t in self.daily_tickets + self.pending_tickets:
            enrich_ticket_fields(t)

        # 工单去重
        original_daily = len(self.daily_tickets)
        self.daily_tickets = dedup_tickets(self.daily_tickets)
        self.dedup_removed = original_daily - len(self.daily_tickets)

        # LLM 分类
        if LLM_ENABLED:
            try:
                from ai_classifier import batch_classify
                batch_classify(self.daily_tickets)
            except Exception as e:
                logger.warning(f"AI分类失败: {e}")

    # ==================== 内容摘要 ====================

    def _get_latest_reply(self, t):
        """获取工单最新回复内容"""
        for entry in reversed(t.get("_log", [])):
            for info in entry.get("info", []):
                c = info.get("content", "")
                if c and "回复" in info.get("title", info.get("titleLang", "")):
                    return c
        return ""

    def _summarize(self, t, max_len=0):
        """
        工单内容摘要。
        优先级：AI摘要 > 最新回复完整内容 > 标题+工单内容
        max_len=0 表示不截断（Web端）。
        """
        # AI摘要优先
        ai_summary = t.get("_ai_summary")
        if ai_summary:
            return ai_summary

        title = t.get("_title", "")
        content = t.get("_content", "")
        latest_reply = self._get_latest_reply(t)

        if latest_reply:
            # 有回复时：标题 + 最新回复完整内容
            summary = f"{title}；[最新回复]{latest_reply}" if title else latest_reply
        else:
            # 无回复时：标题+内容
            parts = []
            if title:
                parts.append(title)
            if content and content != title:
                parts.append(content)
            summary = "；".join(parts)

        if max_len and len(summary) > max_len:
            summary = summary[:max_len - 3] + "..."
        return summary or "（无内容）"

    # ==================== 趋势标记 ====================

    def _trend_mark(self, current, prev_key):
        """生成趋势对比标记"""
        prev = self.trend_data.get(prev_key)
        if prev is None:
            return ""
        diff = current - prev
        if diff > 0:
            return f" (+{diff})"
        elif diff < 0:
            return f" ({diff})"
        return " (持平)"

    # ==================== 格式化 ====================

    def _format_order(self, t, show_amount=False, show_resolved=False):
        """格式化单条工单"""
        oid = t["_id"]
        summary = self._summarize(t, max_len=120)
        creator = t["_creator"] or "未知"
        ct = ts_to_str(t["_create_time"])
        ut = ts_to_str(t["_update_time"])

        line = f"工单号：{oid}，工单内容：{summary}"
        if show_amount:
            amt = t["_recharge"]
            amt_str = f"{amt/10000:.1f}W" if amt >= 10000 else f"{amt:.0f}元"
            line += f"（{amt_str}）"

        line += f"\n    发起人：{creator}，创建时间：{ct}，更新时间：{ut}"

        if show_resolved:
            resolved = t["_status"] in (STATUS_SOLVED, STATUS_CLOSED)
            line += f"，{'已解决' if resolved else '未解决'}"

        return line

    def _format_list(self, orders, **kwargs):
        """格式化工单列表"""
        if not orders:
            return "  无\n"
        lines = []
        for i, t in enumerate(orders, 1):
            lines.append(f"  {i}. {self._format_order(t, **kwargs)}")
        return "\n".join(lines) + "\n"

    # ==================== 七个板块 ====================

    def section_1_major_events(self):
        """一、重大事件报备（批量问题）"""
        title = "一、重大事件报备（批量问题）"
        if not self.session_data:
            return f"{title}\n  无\n"

        keyword_users = defaultdict(set)
        for sess in self.session_data:
            content = sess.get("content", "") or sess.get("message", "") or ""
            visitor = sess.get("visitorName", "") or sess.get("userId", "unknown")
            for category, keywords in ISSUE_KEYWORDS.items():
                if any(kw in content for kw in keywords):
                    keyword_users[category].add(visitor)
                    break

        daily_creators = {t["_creator"] for t in self.daily_tickets if t["_creator"]}
        batch_issues = [
            (cat, len(users), list(users)[:5])
            for cat, users in keyword_users.items()
            if len(users) >= 2 and (users & daily_creators)
        ]

        if not batch_issues:
            return f"{title}\n  无\n"

        lines = [title]
        for cat, count, users in batch_issues:
            lines.append(f"  【{cat}】{count}位用户反馈，涉及用户：{'、'.join(users)}")
        return "\n".join(lines) + "\n"

    def section_2_pending(self):
        """二、待跟进问题总计"""
        title = "二、待跟进问题总计（近30天，受理方：飞鱼科技）"
        orders = [
            t for t in self.pending_tickets
            if HANDLER_DEV_KEYWORD in (t.get("_handler") or "")
        ]
        trend = self._trend_mark(len(orders), "prev_pending_dev_count")
        return f"{title}\n  共 {len(orders)} 条{trend}\n{self._format_list(orders)}"

    def section_3_unvisited(self):
        """三、客服未回访/未跟进问题总计"""
        title = "三、客服未回访/未跟进问题总计（近30天，受理方：工单系统）"
        orders = [
            t for t in self.pending_tickets
            if HANDLER_SYSTEM_KEYWORD in (t.get("_handler") or "")
            or (t.get("_handler") or "").strip() in ("", "空")
        ]
        trend = self._trend_mark(len(orders), "prev_unvisited_count")
        return f"{title}\n  共 {len(orders)} 条{trend}\n{self._format_list(orders)}"

    def section_4_super_r(self):
        """四、超R反馈问题（10w+）"""
        title = "四、超R反馈问题（10w+）"
        orders = sorted(
            [t for t in self.daily_tickets if t["_recharge"] >= SUPER_R_THRESHOLD],
            key=lambda x: x["_recharge"], reverse=True,
        )
        if not orders:
            return f"{title}\n  无\n"
        trend = self._trend_mark(len(orders), "prev_super_r_count")
        return f"{title}\n  共 {len(orders)} 条{trend}\n{self._format_list(orders, show_amount=True)}"

    def section_5_pre_churn(self):
        """五、预流失报备"""
        title = "五、预流失报备"
        orders = [t for t in self.daily_tickets if t.get("_issue_select") == "预流失"]
        if not orders:
            return f"{title}\n  无\n"
        trend = self._trend_mark(len(orders), "prev_pre_churn_count")
        return f"{title}\n  共 {len(orders)} 条{trend}\n{self._format_list(orders)}"

    def section_6_pre_complaint(self):
        """六、预投诉报备"""
        title = "六、预投诉报备"
        orders = [t for t in self.daily_tickets if t.get("_issue_select") == "我要投诉"]
        if not orders:
            return f"{title}\n  无\n"
        trend = self._trend_mark(len(orders), "prev_pre_complaint_count")
        return f"{title}\n  共 {len(orders)} 条{trend}\n{self._format_list(orders)}"

    def section_7_other(self):
        """七、其他VIP用户反馈"""
        title = "七、其他VIP用户反馈"
        classified_selects = {"预流失", "我要投诉"}
        orders = [
            t for t in self.daily_tickets
            if t.get("_issue_select") not in classified_selects
            and t["_recharge"] < SUPER_R_THRESHOLD
        ]
        if not orders:
            return f"{title}\n  无\n"

        grouped = defaultdict(list)
        for t in orders:
            grouped[classify_ticket(t)].append(t)

        lines = [title, f"  共 {len(orders)} 条"]
        for cat, group in grouped.items():
            lines.append(f"\n  【{cat}】({len(group)}条)")
            for i, t in enumerate(group, 1):
                lines.append(f"  {i}. {self._format_order(t, show_resolved=True)}")
        return "\n".join(lines) + "\n"

    # ==================== 统计表 ====================

    def section_stats(self):
        """问题类型统计（文本版）"""
        title = "附：问题类型统计"

        if self.session_data:
            categories, cat_alarm, cat_dev = compute_category_stats(
                self.session_data, self.daily_tickets,
            )
        else:
            categories, cat_alarm, cat_dev = compute_ticket_category_stats(
                self.daily_tickets,
            )

        total_alarm = sum(cat_alarm.values())
        total_dev = sum(cat_dev.values())
        total_agent = total_alarm - total_dev

        lines = [title]
        if not self.session_data:
            lines.append("  [注] 统计基于工单数据（会话监控数据补充中）")
        lines.extend([
            "",
            f"  {'问题类型':<10} {'工单/会话数':>10} {'客服处理':>8} {'运营/研发介入':>12}",
            f"  {'─'*10} {'─'*10} {'─'*8} {'─'*12}",
        ])
        for cat in categories:
            a = cat_alarm.get(cat, 0)
            d = cat_dev.get(cat, 0)
            if a > 0:
                lines.append(f"  {cat:<10} {a:>10} {a-d:>8} {d:>12}")

        lines.extend([
            f"  {'─'*10} {'─'*10} {'─'*8} {'─'*12}",
            f"  {'问题总量':<10} {total_alarm:>10} {total_agent:>8} {total_dev:>12}",
            f"  总会话量：{self.total_sessions}",
        ])

        return "\n".join(lines) + "\n"

    # ==================== 结构化数据（供前端渲染） ====================

    def build_structured(self):
        """生成结构化日报数据（JSON 友好格式，供 Web 前端渲染）"""
        # 分类统计
        if self.session_data:
            categories, cat_alarm, cat_dev = compute_category_stats(
                self.session_data, self.daily_tickets,
            )
        else:
            categories, cat_alarm, cat_dev = compute_ticket_category_stats(
                self.daily_tickets,
            )

        # 超R工单
        super_r_orders = sorted(
            [t for t in self.daily_tickets if t["_recharge"] >= SUPER_R_THRESHOLD],
            key=lambda x: x["_recharge"], reverse=True,
        )

        # 待跟进（飞鱼）
        pending_dev = [
            t for t in self.pending_tickets
            if HANDLER_DEV_KEYWORD in (t.get("_handler") or "")
        ]
        # 未回访
        unvisited = [
            t for t in self.pending_tickets
            if HANDLER_SYSTEM_KEYWORD in (t.get("_handler") or "")
            or (t.get("_handler") or "").strip() in ("", "空")
        ]
        # 预流失
        pre_churn = [t for t in self.daily_tickets if t.get("_issue_select") == "预流失"]
        # 预投诉
        pre_complaint = [t for t in self.daily_tickets if t.get("_issue_select") == "我要投诉"]
        # 其他
        classified_selects = {"预流失", "我要投诉"}
        other_orders = [
            t for t in self.daily_tickets
            if t.get("_issue_select") not in classified_selects
            and t["_recharge"] < SUPER_R_THRESHOLD
        ]
        other_grouped = defaultdict(list)
        for t in other_orders:
            other_grouped[classify_ticket(t)].append(t)

        def _ticket_to_dict(t):
            return {
                "id": t.get("_id", ""),
                "title": t.get("_title", ""),
                "summary": self._summarize(t, max_len=0),
                "creator": t.get("_creator", ""),
                "create_time": ts_to_str(t.get("_create_time", 0)),
                "update_time": ts_to_str(t.get("_update_time", 0)),
                "recharge": t.get("_recharge", 0),
                "recharge_str": f"{t.get('_recharge',0)/10000:.1f}W" if t.get("_recharge", 0) >= 10000 else f"{t.get('_recharge',0):.0f}",
                "status": "已解决" if t.get("_status") in (STATUS_SOLVED, STATUS_CLOSED) else "未解决",
                "handler": t.get("_handler", ""),
                "category": classify_ticket(t),
                "issue_select": t.get("_issue_select", ""),
            }

        return {
            "date": self.report_date.strftime("%Y-%m-%d"),
            "date_cn": self.report_date.strftime("%Y年%m月%d日"),
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "has_session_data": bool(self.session_data),
            "dedup_removed": self.dedup_removed,
            "errors": self.errors,
            "stats": {
                "daily_count": len(self.daily_tickets),
                "pending_count": len(self.pending_tickets),
                "total_sessions": self.total_sessions,
                "super_r_count": len(super_r_orders),
                "pre_churn_count": len(pre_churn),
                "pre_complaint_count": len(pre_complaint),
                "pending_dev_count": len(pending_dev),
                "unvisited_count": len(unvisited),
            },
            "trend": self.trend_data,
            "category_stats": [
                {
                    "name": cat,
                    "total": cat_alarm.get(cat, 0),
                    "agent": cat_alarm.get(cat, 0) - cat_dev.get(cat, 0),
                    "dev": cat_dev.get(cat, 0),
                }
                for cat in categories if cat_alarm.get(cat, 0) > 0
            ],
            "category_stats_total": {
                "total": sum(cat_alarm.values()),
                "agent": sum(cat_alarm.values()) - sum(cat_dev.values()),
                "dev": sum(cat_dev.values()),
            },
            "sections": {
                "super_r": [_ticket_to_dict(t) for t in super_r_orders],
                "pending_dev": [_ticket_to_dict(t) for t in pending_dev],
                "unvisited": [_ticket_to_dict(t) for t in unvisited],
                "pre_churn": [_ticket_to_dict(t) for t in pre_churn],
                "pre_complaint": [_ticket_to_dict(t) for t in pre_complaint],
                "other": {
                    cat: [_ticket_to_dict(t) for t in group]
                    for cat, group in other_grouped.items()
                },
            },
        }

    # ==================== 生成完整日报 ====================

    def build(self):
        """生成完整日报文本"""
        date_str = self.report_date.strftime("%Y年%m月%d日")
        header = (
            f"{'='*50}\n"
            f"  VIP客服日报 - {date_str}\n"
            f"{'='*50}\n"
        )

        # 错误/警告提示
        warnings = []
        if self.errors:
            warnings.append("  [!] 以下数据获取不完整：" + "、".join(self.errors))
        if self.dedup_removed > 0:
            warnings.append(f"  [i] 已自动去除 {self.dedup_removed} 条重复工单")
        warning_section = "\n".join(warnings) + "\n" if warnings else ""

        sections = [
            header,
            warning_section,
            self.section_1_major_events(),
            self.section_2_pending(),
            self.section_3_unvisited(),
            self.section_4_super_r(),
            self.section_5_pre_churn(),
            self.section_6_pre_complaint(),
            self.section_7_other(),
            self.section_stats(),
        ]
        report = "\n".join(sections)
        report += f"\n{'='*50}\n  报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        return report

    # ==================== 文件输出 ====================

    def _output_path(self, ext):
        """生成默认输出文件路径"""
        date_str = self.report_date.strftime("%Y%m%d")
        return Path(OUTPUT_DIR) / f"VIP客服日报_{date_str}.{ext}"

    def save_text(self, filepath=None):
        """保存文本日报"""
        filepath = filepath or self._output_path("txt")
        report = self.build()
        Path(filepath).write_text(report, encoding="utf-8")
        return str(filepath)

    def save_excel(self, filepath=None):
        """保存增强版 Excel 日报（条件格式、图表、冻结表头）"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.utils import get_column_letter

        filepath = filepath or self._output_path("xlsx")
        wb = Workbook()
        ws = wb.active

        _illegal_chars_re = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

        def safe(val):
            return _illegal_chars_re.sub('', val) if isinstance(val, str) else val

        ws.title = "VIP客服日报"

        # ---- 样式 ----
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=10, color="FFFFFF")
        section_font = Font(bold=True, size=11, color="2F5496")
        normal_font = Font(size=10)
        bold_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(bold=True, size=10, color="9C0006")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        yellow_font = Font(size=10, color="9C6500")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(size=10, color="006100")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center")
        wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

        row = 1
        date_str_cn = self.report_date.strftime("%Y年%m月%d日")

        # ---- 标题 ----
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=safe(f"VIP客服日报 - {date_str_cn}")).font = title_font
        row += 1

        # 警告信息
        if self.errors:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=safe(f"[!] 数据获取不完整：{'、'.join(self.errors)}"))
            cell.font = red_font
            cell.fill = red_fill
            row += 1
        if self.dedup_removed > 0:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=1, value=safe(f"[i] 已自动去除 {self.dedup_removed} 条重复工单")).font = normal_font
            row += 1

        row += 1

        # ---- 概览统计卡片 ----
        overview_headers = ["当日工单数", "待跟进工单", "总会话量", "超R工单", "预流失", "预投诉"]
        super_r_count = sum(1 for t in self.daily_tickets if t["_recharge"] >= SUPER_R_THRESHOLD)
        pre_churn_count = sum(1 for t in self.daily_tickets if t.get("_issue_select") == "预流失")
        pre_complaint_count = sum(1 for t in self.daily_tickets if t.get("_issue_select") == "我要投诉")
        overview_values = [
            len(self.daily_tickets), len(self.pending_tickets), self.total_sessions,
            super_r_count, pre_churn_count, pre_complaint_count,
        ]
        for col, h in enumerate(overview_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center
        row += 1
        for col, v in enumerate(overview_values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(bold=True, size=14, color="2F5496")
            cell.border = thin_border
            cell.alignment = center
        # 趋势行
        if self.trend_data:
            row += 1
            trend_keys = ["prev_daily_count", "prev_pending_count", "prev_total_sessions",
                          "prev_super_r_count", "prev_pre_churn_count", "prev_pre_complaint_count"]
            for col, (val, key) in enumerate(zip(overview_values, trend_keys), 1):
                prev = self.trend_data.get(key)
                if prev is not None:
                    diff = val - prev
                    mark = f"+{diff}" if diff > 0 else str(diff) if diff < 0 else "持平"
                    cell = ws.cell(row=row, column=col, value=safe(f"较昨日 {mark}"))
                    cell.alignment = center
                    if diff > 0:
                        cell.font = red_font
                    elif diff < 0:
                        cell.font = green_font
                    else:
                        cell.font = normal_font
        row += 2

        # ---- 统计表 ----
        if self.session_data:
            categories, cat_alarm, cat_dev = compute_category_stats(
                self.session_data, self.daily_tickets,
            )
        else:
            categories, cat_alarm, cat_dev = compute_ticket_category_stats(
                self.daily_tickets,
            )

        stat_headers = ["问题类型", "工单/会话数", "客服处理", "运营/研发介入"]
        for col, h in enumerate(stat_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center
        row += 1

        chart_start_row = row
        for cat in categories:
            a = cat_alarm.get(cat, 0)
            d = cat_dev.get(cat, 0)
            if a == 0:
                continue
            vals = [cat, a, a - d, d]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=safe(v))
                cell.font = normal_font
                cell.border = thin_border
                if col > 1:
                    cell.alignment = center
            row += 1
        chart_end_row = row - 1

        # 合计行
        total_a = sum(cat_alarm.values())
        total_d = sum(cat_dev.values())
        for col, v in enumerate(["合计", total_a, total_a - total_d, total_d], 1):
            cell = ws.cell(row=row, column=col, value=safe(v))
            cell.font = bold_font
            cell.border = thin_border
            cell.fill = light_fill
            if col > 1:
                cell.alignment = center
        row += 1

        ws.cell(row=row, column=1, value=safe(f"总会话量：{self.total_sessions}")).font = bold_font
        row += 1

        # ---- 饼图：问题类型分布 ----
        if chart_end_row >= chart_start_row:
            pie = PieChart()
            pie.title = "问题类型分布"
            pie.width = 18
            pie.height = 12
            labels = Reference(ws, min_col=1, min_row=chart_start_row, max_row=chart_end_row)
            data = Reference(ws, min_col=2, min_row=chart_start_row - 1, max_row=chart_end_row)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.style = 10
            ws.add_chart(pie, f"F{chart_start_row}")

        row += 16  # 为图表留空间

        # ---- 七个板块（详细工单表格） ----

        def _write_section_table(ws, row, section_title, orders,
                                 show_amount=False, show_resolved=False):
            """写一个板块的工单表格"""
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=safe(section_title))
            cell.font = section_font
            row += 1

            if not orders:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                ws.cell(row=row, column=1, value="无").font = normal_font
                return row + 1

            # 表头
            cols = ["序号", "发起人", "工单内容", "创建时间", "更新时间", "受理人"]
            if show_amount:
                cols.append("累充金额")
            if show_resolved:
                cols.append("状态")
            for col, h in enumerate(cols, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center
            row += 1

            for i, t in enumerate(orders, 1):
                values = [
                    i,
                    t.get("_creator", ""),
                    self._summarize(t, max_len=200),
                    ts_to_str(t.get("_create_time", 0)),
                    ts_to_str(t.get("_update_time", 0)),
                    t.get("_handler", ""),
                ]
                if show_amount:
                    amt = t["_recharge"]
                    values.append(f"{amt/10000:.1f}W" if amt >= 10000 else f"{amt:.0f}")
                if show_resolved:
                    resolved = t["_status"] in (STATUS_SOLVED, STATUS_CLOSED)
                    values.append("已解决" if resolved else "未解决")

                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=safe(v))
                    cell.font = normal_font
                    cell.border = thin_border
                    if col == 3:  # 工单内容列自动换行
                        cell.alignment = wrap
                    else:
                        cell.alignment = Alignment(vertical="center")

                # 条件格式：超R行红色高亮
                if show_amount and t["_recharge"] >= SUPER_R_THRESHOLD:
                    for col in range(1, len(values) + 1):
                        ws.cell(row=row, column=col).fill = red_fill
                        ws.cell(row=row, column=col).font = red_font

                # 条件格式：未解决标黄
                if show_resolved and t["_status"] not in (STATUS_SOLVED, STATUS_CLOSED):
                    for col in range(1, len(values) + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill
                        if col == len(values):
                            ws.cell(row=row, column=col).font = yellow_font

                # 条件格式：已解决标绿
                if show_resolved and t["_status"] in (STATUS_SOLVED, STATUS_CLOSED):
                    status_col = len(values)
                    ws.cell(row=row, column=status_col).fill = green_fill
                    ws.cell(row=row, column=status_col).font = green_font

                row += 1

            return row + 1

        # 各板块
        # 二、待跟进
        pending_dev = [t for t in self.pending_tickets if HANDLER_DEV_KEYWORD in (t.get("_handler") or "")]
        row = _write_section_table(ws, row, f"二、待跟进问题总计 (共{len(pending_dev)}条)", pending_dev)

        # 三、未回访
        unvisited = [
            t for t in self.pending_tickets
            if HANDLER_SYSTEM_KEYWORD in (t.get("_handler") or "")
            or (t.get("_handler") or "").strip() in ("", "空")
        ]
        row = _write_section_table(ws, row, f"三、客服未回访/未跟进 (共{len(unvisited)}条)", unvisited)

        # 四、超R
        super_r = sorted(
            [t for t in self.daily_tickets if t["_recharge"] >= SUPER_R_THRESHOLD],
            key=lambda x: x["_recharge"], reverse=True,
        )
        row = _write_section_table(ws, row, f"四、超R反馈问题 (共{len(super_r)}条)", super_r, show_amount=True)

        # 五、预流失
        pre_churn = [t for t in self.daily_tickets if t.get("_issue_select") == "预流失"]
        row = _write_section_table(ws, row, f"五、预流失报备 (共{len(pre_churn)}条)", pre_churn)

        # 六、预投诉
        pre_complaint = [t for t in self.daily_tickets if t.get("_issue_select") == "我要投诉"]
        row = _write_section_table(ws, row, f"六、预投诉报备 (共{len(pre_complaint)}条)", pre_complaint, show_resolved=True)

        # 七、其他
        classified_selects = {"预流失", "我要投诉"}
        other_orders = [
            t for t in self.daily_tickets
            if t.get("_issue_select") not in classified_selects
            and t["_recharge"] < SUPER_R_THRESHOLD
        ]
        row = _write_section_table(ws, row, f"七、其他VIP用户反馈 (共{len(other_orders)}条)", other_orders, show_resolved=True)

        # ---- 列宽 ----
        ws.column_dimensions["A"].width = 6   # 序号
        ws.column_dimensions["B"].width = 14  # 发起人
        ws.column_dimensions["C"].width = 55  # 工单内容
        ws.column_dimensions["D"].width = 18  # 创建时间
        ws.column_dimensions["E"].width = 18  # 更新时间
        ws.column_dimensions["F"].width = 16  # 受理人
        ws.column_dimensions["G"].width = 12  # 累充金额/状态
        ws.column_dimensions["H"].width = 10

        # ---- 冻结表头 ----
        ws.freeze_panes = "A4"

        # ---- 超R金额柱状图（独立Sheet）----
        if super_r:
            ws2 = wb.create_sheet("超R金额分布")
            ws2.cell(row=1, column=1, value="玩家").font = header_font
            ws2.cell(row=1, column=1).fill = header_fill
            ws2.cell(row=1, column=2, value="累充金额(万)").font = header_font
            ws2.cell(row=1, column=2).fill = header_fill
            for i, t in enumerate(super_r, 2):
                ws2.cell(row=i, column=1, value=safe(t.get("_creator", ""))).font = normal_font
                ws2.cell(row=i, column=2, value=round(t["_recharge"] / 10000, 1)).font = normal_font
            bar = BarChart()
            bar.title = "超R玩家累充金额分布"
            bar.y_axis.title = "金额（万元）"
            bar.x_axis.title = "玩家"
            bar.width = 24
            bar.height = 14
            cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(super_r))
            vals = Reference(ws2, min_col=2, min_row=1, max_row=1 + len(super_r))
            bar.add_data(vals, titles_from_data=True)
            bar.set_categories(cats)
            bar.style = 10
            ws2.add_chart(bar, "D1")
            ws2.column_dimensions["A"].width = 20
            ws2.column_dimensions["B"].width = 15

        # 写入文件
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=Path(filepath).parent)
        try:
            os.close(tmp_fd)
            wb.save(tmp_path)
            target = Path(filepath)
            if target.exists():
                try:
                    target.unlink()
                except PermissionError:
                    ts = datetime.now().strftime("%H%M%S")
                    filepath = str(target.with_stem(target.stem + f"_{ts}"))
            Path(tmp_path).replace(filepath)
        except Exception:
            Path(tmp_path).unlink(missing_ok=True)
            raise

        logger.info(f"Excel日报已保存: {filepath}")
        return str(filepath)

    def save_pdf(self, filepath=None):
        """保存 PDF 日报"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        filepath = filepath or self._output_path("pdf")

        # 注册中文字体
        font_registered = False
        font_paths = [
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/simsun.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont("ChineseFont", fp))
                    font_registered = True
                    break
                except Exception:
                    continue

        if not font_registered:
            logger.warning("未找到中文字体，PDF可能无法正确显示中文")
            font_name = "Helvetica"
        else:
            font_name = "ChineseFont"

        doc = SimpleDocTemplate(
            str(filepath), pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ChTitle', parent=styles['Title'],
            fontName=font_name, fontSize=16, spaceAfter=12,
        )
        heading_style = ParagraphStyle(
            'ChHeading', parent=styles['Heading2'],
            fontName=font_name, fontSize=12, spaceAfter=6,
            textColor=colors.HexColor("#2F5496"),
        )
        body_style = ParagraphStyle(
            'ChBody', parent=styles['Normal'],
            fontName=font_name, fontSize=9, leading=14,
            spaceAfter=4,
        )

        elements = []
        date_str_cn = self.report_date.strftime("%Y年%m月%d日")
        elements.append(Paragraph(f"VIP客服日报 - {date_str_cn}", title_style))
        elements.append(Spacer(1, 6*mm))

        # 概览
        overview_data = [
            ["当日工单", "待跟进", "总会话量", "超R工单", "预流失", "预投诉"],
            [
                str(len(self.daily_tickets)),
                str(len(self.pending_tickets)),
                str(self.total_sessions),
                str(sum(1 for t in self.daily_tickets if t["_recharge"] >= SUPER_R_THRESHOLD)),
                str(sum(1 for t in self.daily_tickets if t.get("_issue_select") == "预流失")),
                str(sum(1 for t in self.daily_tickets if t.get("_issue_select") == "我要投诉")),
            ],
        ]
        overview_table = Table(overview_data, colWidths=[28*mm]*6)
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, 1), 14),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(overview_table)
        elements.append(Spacer(1, 8*mm))

        # 各板块文本
        report_text = self.build()
        for line in report_text.split("\n"):
            line = line.rstrip()
            if not line:
                elements.append(Spacer(1, 2*mm))
                continue
            if line.startswith("="):
                continue
            if line.lstrip().startswith(("一、", "二、", "三、", "四、", "五、", "六、", "七、", "附：")):
                elements.append(Paragraph(line.strip(), heading_style))
            else:
                # 转义 HTML 特殊字符
                safe_line = (line.replace("&", "&amp;")
                             .replace("<", "&lt;")
                             .replace(">", "&gt;"))
                elements.append(Paragraph(safe_line, body_style))

        # 页脚
        elements.append(Spacer(1, 10*mm))
        footer_style = ParagraphStyle(
            'Footer', parent=styles['Normal'],
            fontName=font_name, fontSize=8, textColor=colors.grey,
        )
        elements.append(Paragraph(
            f"报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            footer_style,
        ))

        doc.build(elements)
        logger.info(f"PDF日报已保存: {filepath}")
        return str(filepath)
